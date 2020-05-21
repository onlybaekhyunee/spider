import requests
import bs4
import re
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,Alignment,NamedStyle,Border,Side,PatternFill
import datetime
from functools import reduce

class Spider():
    def __init__(self,url):
        self.url = url

    def __open_url(self):
        """ 根据url请求数据 """
        # 设置访问方式为google浏览器，user_agent识别为python会被反爬阻止
        user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1 Safari/605.1.15'
        headers = {'user-agent' : user_agent}
        res = requests.get(self.url,headers = headers)
        return res

    def __fetch_content(self,res):
        """ 将request获取的内容转为BeautifulSoup """
        # 第二个参数为对应的解析器
        soup = bs4.BeautifulSoup(res.text,'html.parser')
        return soup 

    def __analysis(self,soup):
        """ BeautifulSoup数据分析 """
        targets = soup.find_all('div',class_='info')
        movies = []
        for each in targets:
            # 获取电影名称
            names = each.find('div',class_='hd').a.get_text(strip=True) 
            name = '\n'.join(names.split('/')[0:2])

            bd = each.find('div',class_='bd')
            # 获取评分 
            rating = bd.find('div',class_='star').find('span',class_='rating_num').get_text()
            # 获取导演演员表
            info = bd.find('p').get_text(strip=True)
            s = re.findall('导演:(.*?)主',info)
            director = s[0].strip() if len(s)>0 else ''
            s = re.findall('主演:(.*?)\d',info)
            actor = s[0].strip() if len(s)>0 else ''
            # 获取上映时间
            s = re.findall('(\d+)',info)
            year = s[0].strip() if len(s)>0 else ''
            # 获取国家
            s = re.findall('\d+\s*/(\D*?)/',info)
            country = s[0].strip() if len(s)>0 else ''
            # 获取简介
            tag = bd.find('p',class_='quote')
            quote = '' if tag is None else tag.get_text(strip=True)
            movie = {
                'name':name,
                'rating':rating,
                'director':director,
                'actor':actor,
                'year':year,
                'country':country,
                'quote':quote
            }
            movies.append(movie)       
        return movies
        
    def __find_depth(self,soup):
        """ 获取总页数 """       
        depth = soup.find('span',class_='next').previous_sibling.previous_sibling.text
        return int(depth)

    def __sort(self,movies,sign):
        """ 数据排序，使用的内置函数sorted """
        movies = sorted(movies,key=lambda x: x[sign],reverse=True)
        return movies

    def __saveExc(self,movies,filename,sheetname):
        """ 将数据存储至Excel """
        wb = Workbook()
        ws = wb.create_sheet(index=0,title=sheetname)
        # 设置主标题
        nt = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A1'] = '电影排行榜250 ('+nt+')'
        # 设置列标题
        ws.append(['序号','电影名称','评分','导演','演员','上映年份','国家','简介'])
        num = 1
        for m in movies:
            l = [num]
            l.extend([value for key,value in m.items()])
            ws.append(l)
            num += 1
        try:           
            wb.save(filename)
        except Exception as e:
            print(e)
            return False
        else:
            return True

    def __format_Exc(self,filename,sheetname):
        """ 格式化Excell """
        wb = load_workbook(filename)
        ws = wb[sheetname]
        # 合并标题行单元格
        ws.merge_cells('A1:H1') 
        # 设置标题行样式  
        top_left_cell = ws['A1']
        top_left_cell.font = Font(name='微软雅黑',b=True,size=16)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws['H1'].border = Border(right=Side(border_style='thin'))
        top_left_cell.fill = PatternFill("solid", fgColor="BCEE68")
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 20
        thin=Side(border_style='thin')
        for col in range(1,9):
            ws.cell(row=2,column=col).font = Font(name='微软雅黑',b=True,size=12)
            ws.cell(row=2,column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=2,column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws.cell(row=2,column=col).fill = PatternFill("solid", fgColor="BCEE68")
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 80
        # 正文通用样式
        fontformat = NamedStyle(name='fontformat')
        fontformat.font = Font(name='微软雅黑',size=12)
        fontformat.alignment = Alignment(horizontal="left", vertical="center",wrap_text=True)       
        fontformat.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wb.add_named_style(fontformat)
        fill1 = PatternFill("solid", fgColor="D1EEEE")
        fill2 = PatternFill("solid", fgColor="F4F4F4")
        # 遍历表格，设置样式
        for row in ws.iter_rows(min_col=1,min_row=3,max_col=ws.max_column,max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 50
            ws[row[2].coordinate].value = float(ws[row[2].coordinate].value)
            ws[row[5].coordinate].value = int(ws[row[5].coordinate].value)
            fill = fill2 if int(row[0].row)%2 == 0 else fill1
            for cell in row:
                cell.style = fontformat  
                cell.fill = fill      
        ws.freeze_panes = 'A3'
        try:           
            wb.save(filename)
        except Exception as e:
            print(e)
            return False
        else:
            return True

    def go(self):
        """ 入口文件 """
        res = self.__open_url()
        if res.status_code == 200:
            # 从服务器获取的数据不为空
            soup = self.__fetch_content(res)
            depth = self.__find_depth(soup)
            url = self.url
            movies = []
            for page_num in range(depth):
                self.url = url+'?start='+str(25*page_num)
                res = self.__open_url()
                soup = self.__fetch_content(res)
                movies.extend(self.__analysis(soup))
            movies = self.__sort(movies,'rating')
            succes = self.__saveExc(movies,'MoviesTop250.xlsx','MoviesTop250')
            if succes:
                self.__format_Exc('MoviesTop250.xlsx','MoviesTop250')
       

url = 'https://movie.douban.com/top250'
s = Spider(url)
s.go()


